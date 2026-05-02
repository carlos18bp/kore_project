"""Import exercises from Excel (EN) or CSV (ES) into the Exercise model.

Usage:
    python manage.py import_exercises                          # English Excel
    python manage.py import_exercises --path file.xlsx        # Custom Excel
    python manage.py import_exercises --csv file.csv          # Spanish CSV (update by URL)

The command is idempotent: re-running updates existing records.
English Excel uses name as the unique key; Spanish CSV uses youtube_url.
"""

import csv
import os

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from core_app.models.exercise import Exercise


# ── Spanish → English vocabulary maps ─────────────────────────────────────────

# Maps any variant → canonical Spanish value stored in DB.
# The canonical values must match PATTERN_SEQUENCE_BY_GOAL in program_generator.py.
_PATTERN_CANONICAL = {
    'sentadilla': 'Sentadilla',
    'empuje': 'Empuje', 'empujar': 'Empuje',
    'jalar': 'Jalar', 'jalón': 'Jalar', 'tirar': 'Jalar',
    'doblar': 'Doblar', 'doblez': 'Doblar', 'doblarse': 'Doblar',
    'flexión': 'Doblar', 'flexionar': 'Doblar',
    'llevar': 'Llevar',
    'core': 'Núcleo', 'núcleo': 'Núcleo',
    'rotación': 'Rotación',
    'locomoción': 'Locomoción',
    'explosividad': 'Pliométrico', 'pliométrico': 'Pliométrico', 'plyométrico': 'Pliométrico',
    'complejo': 'Complejo',
    'una pierna': 'Una pierna', 'pierna única': 'Una pierna',
    'monostructural': 'Monostructural',
}

# Maps any variant → canonical Spanish type stored in DB.
_TYPE_CANONICAL = {
    'peso corporal': 'Peso corporal', 'propio peso': 'Peso corporal',
    'carga externa': 'Carga externa',
    'estabilidad': 'Estabilidad',
    'explosividad': 'Explosividad',
    'correctivo': 'Correctivo',
    'movilidad': 'Movilidad',
}

# English → English for _fitness_level / _goal_tags logic (unchanged)
_PATTERN_ES = {
    'sentadilla': 'Squat',
    'empuje': 'Push', 'empujar': 'Push',
    'jalar': 'Pull', 'jalón': 'Pull', 'tirar': 'Pull',
    'doblar': 'Bend', 'doblez': 'Bend', 'doblarse': 'Bend',
    'llevar': 'Carry',
    'core': 'Core', 'núcleo': 'Core',
    'rotación': 'Rotation',
    'locomoción': 'Locomotion',
    'flexión': 'Hinge', 'flexionar': 'Hinge',
    'explosividad': 'Explosive', 'pliométrico': 'Plyometric', 'plyométrico': 'Plyometric',
    'complejo': 'Complex',
    'una pierna': 'Single Leg', 'pierna única': 'Single Leg',
    'anti-extensión': 'Anti-Extension', 'anti-rotación': 'Anti-Rotation',
    'monostructural': 'Monostructural',
}

_TYPE_ES = {
    'peso corporal': 'Bodyweight', 'propio peso': 'Bodyweight',
    'carga externa': 'External Loading',
    'estabilidad': 'Stability',
    'explosividad': 'Explosiveness',
    'correctivo': 'Corrective',
    'movilidad': 'Mobility',
}

_IMPLEMENT_ES = {
    'barra': 'Barbell', 'barra olímpica': 'Barbell', 'barra axel': 'Axle Bar',
    'barra trap': 'Trap Bar', 'barra de trampa': 'Trap Bar',
    'barra de seguridad': 'Safety Bar', 'barra ez': 'EZ Bar',
    'barra landmine': 'Landmine Barbell', 'landmine': 'Landmine Barbell',
    'mancuerna': 'Dumbbell', 'mancuernas': 'Dumbbell',
    'pesa rusa': 'Kettlebell', 'pesas rusas': 'Kettlebell', 'kettlebell': 'Kettlebell',
    'banda': 'Band', 'bandas': 'Band', 'banda de resistencia': 'Band',
    'bandas de resistencia': 'Band', 'bandita': 'Band',
    'cable': 'Cable', 'cables': 'Cable', 'polea': 'Cable', 'poleas': 'Cable',
    'máquina': 'Machine',
    'balón medicinal': 'Medicine Ball', 'balón': 'Medicine Ball',
    'balón de estabilidad': 'Swiss Ball',
    'anillas': 'Rings', 'anillos': 'Rings', 'aros': 'Rings',
    'gHD': 'GHD', 'ghd': 'GHD',
    'trx': 'TRX',
    'trineo': 'Sled', 'sled': 'Sled',
    'banco': 'Bench', 'caja': 'Plyo Box', 'step': 'Step',
    'ninguno': 'None', 'suelo': 'None', 'pared': 'Wall', 'colchoneta': 'None',
    'cuerda': 'Rope',
    'rack': 'Rack', 'paralelas': 'Parallettes',
    'barra de dominadas': 'Pull-Up Bar',
    'rodillo de espuma': 'Foam Roller',
}

_STATUS_ACTIVE_ES = {'hecho', 'realizado', 'filmado', 'done', 'white label'}


def _normalize_es(spa_val: str, mapping: dict) -> str:
    return mapping.get(spa_val.lower().strip(), spa_val.strip())


# Implements considered "light" → level 2 when External Loading
_LIGHT_IMPLEMENTS = {
    'Band', 'Swiss Ball', 'Medicine Ball', 'Foam Roller', 'Abmat',
    'Slider', 'None', 'PVC Pipe', 'Towel', 'Wall', 'Parallettes',
    'Jump Rope', 'Ladder',
}

# Implements requiring high strength base → level 4
_HEAVY_IMPLEMENTS = {
    'Barbell', 'Trap Bar', 'Safety Bar', 'Axle Bar', 'Cambered Bar',
    'Yoke Bar', 'Landmine Barbell', 'Atlas Stone', 'D-Ball', 'Sandbag', 'Sled',
    'Nordic Bench', 'Stall Bars', 'GHD', 'Reverse Hyper', 'Rower',
    'Airdyne', 'Bike Erg', 'Ski Erg', 'Rope', 'Rings', 'Weighted Belt',
}

_ALL_GOALS = ['fat_loss', 'muscle_gain', 'rehab', 'general_health', 'sports_performance']


def _fitness_level(exercise_type: str, pattern: str, main_implement: str) -> int:
    t = (exercise_type or '').strip()
    p = (pattern or '').strip()
    m = (main_implement or '').strip()

    if t == 'Stability':
        return 1

    if t == 'Bodyweight':
        if p == 'Core':
            return 1
        if p in ('Push', 'Pull', 'Single Leg', 'Squat', 'Bend', 'Locomotion'):
            return 2
        return 3  # Complex, Plyometric

    if t == 'Explosiveness':
        if p in ('Monostructural', 'Plyometric', 'Locomotion'):
            return 3
        if p == 'Complex':
            return 5
        return 4  # Push, Pull, Squat, Bend, Single Leg

    # External Loading
    if m in _LIGHT_IMPLEMENTS:
        base = 2
    elif m in _HEAVY_IMPLEMENTS:
        base = 4
    else:
        base = 3  # Dumbbell, Kettlebell, Cable, Machine, EZ Bar, etc.

    if p == 'Complex':
        base = min(base + 1, 5)

    return base


def _goal_tags(exercise_type: str, pattern: str) -> list:
    t = (exercise_type or '').strip()
    p = (pattern or '').strip()

    tags = {'general_health'}

    if t == 'Stability':
        tags.update(['rehab'])
    if p == 'Core':
        tags.update(['rehab'])
    if p == 'Monostructural':
        tags.update(['fat_loss'])
    if p == 'Locomotion':
        tags.update(['fat_loss'])
    if p == 'Plyometric':
        tags.update(['fat_loss', 'sports_performance'])
    if t == 'Explosiveness':
        tags.update(['sports_performance', 'fat_loss'])
    if p in ('Push', 'Pull', 'Bend', 'Squat', 'Complex') and t == 'External Loading':
        tags.update(['muscle_gain', 'sports_performance'])
    if p in ('Push', 'Pull', 'Bend', 'Squat') and t == 'Bodyweight':
        tags.update(['muscle_gain'])
    if p == 'Single Leg':
        tags.update(['rehab', 'sports_performance'])

    return sorted(tags)


class Command(BaseCommand):
    help = 'Import exercises from tier2_exercises_with_links.xlsx (EN) or tier2_exercises_spa.csv (ES)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--path',
            default=None,
            help='Path to the Excel file (defaults to <project_root>/tier2_exercises_with_links.xlsx)',
        )
        parser.add_argument(
            '--csv',
            default=None,
            help='Path to the Spanish CSV file (uses youtube_url as unique key)',
        )

    def handle(self, *args, **options):
        if options['csv']:
            self._import_csv(options['csv'])
        else:
            self._import_excel(options['path'])

    def _import_excel(self, path):
        if path is None:
            base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(
                os.path.abspath(__file__)
            ))))
            path = os.path.join(base, 'tier2_exercises_with_links.xlsx')

        if not os.path.exists(path):
            raise CommandError(f'File not found: {path}')

        self.stdout.write(f'Loading Excel: {path}...')
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        created = updated = skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = (row[0] or '').strip()
            if not name:
                skipped += 1
                continue

            status      = str(row[1] or '').strip()
            pattern     = str(row[2] or '').strip()
            ex_type     = str(row[3] or '').strip()
            implement   = str(row[4] or '').strip()
            primary     = str(row[5] or '').strip()
            secondary   = str(row[6] or '').strip()
            plane       = str(row[7] or '').strip()
            explanation = str(row[8] or '').strip()
            youtube_url = str(row[9] or '').strip()

            level      = _fitness_level(ex_type, pattern, implement)
            corrective = (ex_type == 'Stability')
            tags       = _goal_tags(ex_type, pattern)

            obj, was_created = Exercise.objects.update_or_create(
                name=name,
                defaults=dict(
                    pattern=pattern,
                    exercise_type=ex_type,
                    main_implement=implement,
                    primary_muscles=primary,
                    secondary_muscles=secondary,
                    plane=plane,
                    explanation=explanation,
                    youtube_url=youtube_url if youtube_url.startswith('http') else '',
                    fitness_level_min=level,
                    is_corrective=corrective,
                    goal_tags=tags,
                    is_active=(status in ('done', 'white label')),
                ),
            )
            if was_created:
                created += 1
            else:
                updated += 1

        wb.close()
        self.stdout.write(self.style.SUCCESS(
            f'Done — created: {created}, updated: {updated}, skipped (no name): {skipped}'
        ))

    def _import_csv(self, path):
        if not os.path.exists(path):
            raise CommandError(f'File not found: {path}')

        self.stdout.write(f'Loading Spanish CSV: {path}...')
        created = updated = skipped = 0

        with open(path, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                name        = row.get('Ejercicio', '').strip()
                status      = row.get('Estado', '').strip()
                pattern_es  = row.get('Patrón', '').strip()
                type_es     = row.get('Tipo', '').strip()
                implement_es = row.get('Implementación principal', '').strip()
                primary     = row.get('Músculos primarios trabajados', '').strip()
                secondary   = row.get('Músculos secundarios trabajados', '').strip()
                plane       = row.get('Plano', '').strip()
                explanation = row.get('Explicación del ejercicio', '').strip()
                youtube_url = row.get('URL de YouTube', '').strip()

                if not name or not youtube_url.startswith('http'):
                    skipped += 1
                    continue

                # English values for _fitness_level / _goal_tags logic
                pattern_eng  = _normalize_es(pattern_es, _PATTERN_ES)
                ex_type_eng  = _normalize_es(type_es, _TYPE_ES)
                implement_eng = _normalize_es(implement_es, _IMPLEMENT_ES)

                # Canonical Spanish values for DB storage (match program_generator)
                pattern_db  = _normalize_es(pattern_es, _PATTERN_CANONICAL) or pattern_es
                ex_type_db  = _normalize_es(type_es, _TYPE_CANONICAL) or type_es

                level      = _fitness_level(ex_type_eng, pattern_eng, implement_eng)
                corrective = (ex_type_eng in ('Stability', 'Corrective', 'Mobility'))
                tags       = _goal_tags(ex_type_eng, pattern_eng)

                fields = dict(
                    name=name,
                    pattern=pattern_db,
                    exercise_type=ex_type_db,
                    main_implement=implement_es,
                    primary_muscles=primary,
                    secondary_muscles=secondary,
                    plane=plane,
                    explanation=explanation,
                    youtube_url=youtube_url,
                    fitness_level_min=level,
                    is_corrective=corrective,
                    goal_tags=tags,
                    is_active=(status.lower() in _STATUS_ACTIVE_ES),
                )

                # Look up by URL first (previously imported English record)
                obj = Exercise.objects.filter(youtube_url=youtube_url).first()
                if obj is None:
                    # Fall back: exact name match
                    obj = Exercise.objects.filter(name=name).first()
                    if obj:
                        obj.youtube_url = youtube_url

                vid_id = youtube_url.split('v=')[-1][:8]

                if obj is not None:
                    for k, v in fields.items():
                        setattr(obj, k, v)
                    try:
                        obj.save()
                    except Exception:
                        # Name already taken by a different record — use video ID suffix
                        obj.name = f"{name} ({vid_id})"
                        obj.save()
                    updated += 1
                else:
                    # Avoid name collision: append video ID if name already taken
                    if Exercise.objects.filter(name=name).exists():
                        fields['name'] = f"{name} ({vid_id})"
                    Exercise.objects.create(**fields)
                    created += 1

        self.stdout.write(self.style.SUCCESS(
            f'Done — created: {created}, updated: {updated}, skipped: {skipped}'
        ))
